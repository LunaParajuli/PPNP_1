from odoo import models
import io
import base64
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D

from PIL import Image as PILImage


_logger = logging.getLogger(__name__)


class SalespersonReportExcel(models.AbstractModel):
    _name = 'report.ppnpl_report_wizard.report_salesperson_excel'
    _inherit = 'report.ppnpl_report_wizard.report_salesperson_template'
    _description = 'Salesperson Report Excel'

    def _generate_excel(self, values):

        wb = Workbook()
        ws = wb.active
        ws.title = "Salesperson Report"

        company = self.env.company
        image_buffers = []

        # =========================
        # STYLES
        # =========================
        header_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        header_fill = PatternFill('solid', start_color='2E75B6')
        title_font = Font(name='Arial', bold=True, size=14)
        bold_font = Font(name='Arial', bold=True, size=10)
        normal_font = Font(name='Arial', size=10)

        center = Alignment(horizontal='center', vertical='center')
        left = Alignment(horizontal='left', vertical='center')
        right = Alignment(horizontal='right', vertical='center')

        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        total_fill = PatternFill('solid', start_color='D9E1F2')

        # =========================
        # COLUMN WIDTHS
        # =========================
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15

        # =========================
        # COMPANY HEADER
        # =========================

        # --- Company Logo ---
        if company.logo:
            try:
                logo_data = base64.b64decode(company.logo)
                logo_io = io.BytesIO(logo_data)
                pil_logo = PILImage.open(logo_io)

                pil_logo.thumbnail((140, 70), PILImage.LANCZOS)

                buffer_logo = io.BytesIO()
                pil_logo.save(buffer_logo, format='PNG')
                buffer_logo.seek(0)
                image_buffers.append(buffer_logo)

                xl_logo = XLImage(buffer_logo)
                xl_logo.width = 140
                xl_logo.height = 70

                ws.add_image(xl_logo, "A1")

            except Exception as e:
                _logger.error("Logo Error: %s", e)

        # --- Company Name ---
        ws.merge_cells('C1:G1')
        ws['C1'] = company.name
        ws['C1'].font = title_font
        ws['C1'].alignment = left

        # --- Address ---
        ws.merge_cells('C2:G2')
        ws['C2'] = f"{company.street or ''}, {company.city or ''}, {company.country_id.name or ''}"
        ws['C2'].font = normal_font
        ws['C2'].alignment = left

        # Spacer
        ws.append([])

        # =========================
        # REPORT TITLE
        # =========================
        title_row = ws.max_row + 1
        ws.merge_cells(f'A{title_row}:G{title_row}')
        ws[f'A{title_row}'] = 'Salesperson Performance Report'
        ws[f'A{title_row}'].font = title_font
        ws[f'A{title_row}'].alignment = center

        # Report Info
        ws.append([f"Salesperson: {values['salesperson_name']}"])
        ws.append([f"Period: {values['date_from']} to {values['date_to']}"])
        ws.append([])

        # =========================
        # TABLE HEADER
        # =========================
        header_row = ws.max_row + 1

        headers = [
            'S.No',
            'Product Image',
            'Product',
            'Qty',
            'Unit Price (Rs)',
            'Tax',
            'Amount (Rs)'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        # =========================
        # DATA ROWS
        # =========================

        COL_B_WIDTH_PX = 160
        ROW_HEIGHT_PX = 80

        for idx, line in enumerate(values['docs'], start=1):

            row = ws.max_row + 1
            ws.row_dimensions[row].height = 75

            row_data = [
                idx,
                '',
                line.product_id.name,
                line.product_uom_qty,
                line.price_unit,
                sum(line.tax_ids.mapped('amount')),
                line.price_total,
            ]

            for col, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = normal_font
                cell.border = border

                if col in (4, 5, 6, 7):
                    cell.alignment = right
                    cell.number_format = '#,##0.00'
                else:
                    cell.alignment = left

            # =========================
            # PRODUCT IMAGE
            # =========================
            if line.product_id.image_128:
                try:
                    img_data = base64.b64decode(line.product_id.image_128)
                    img_io = io.BytesIO(img_data)
                    pil_img = PILImage.open(img_io)

                    if pil_img.mode in ('RGBA', 'P'):
                        pil_img = pil_img.convert('RGB')

                    img_w, img_h = 70, 70
                    pil_img.thumbnail((img_w, img_h), PILImage.LANCZOS)

                    buffer_img = io.BytesIO()
                    pil_img.save(buffer_img, format='PNG')
                    buffer_img.seek(0)
                    image_buffers.append(buffer_img)

                    xl_img = XLImage(buffer_img)

                    # Convert pixel to EMU
                    width_emu = img_w * 9525
                    height_emu = img_h * 9525

                    left_offset = int((COL_B_WIDTH_PX - img_w) / 2) * 9525
                    top_offset = int((ROW_HEIGHT_PX - img_h) / 2) * 9525

                    marker = AnchorMarker(
                        col=1,
                        colOff=left_offset,
                        row=row - 1,
                        rowOff=top_offset
                    )

                    size = XDRPositiveSize2D(width_emu, height_emu)

                    xl_img.anchor = OneCellAnchor(_from=marker, ext=size)

                    ws.add_image(xl_img)

                except Exception as e:
                    _logger.error("Product Image Error: %s", e)

        # =========================
        # FOOTER SECTION
        # =========================
        total_row = ws.max_row + 1

        ws.merge_cells(f'A{total_row}:F{total_row}')
        total_label = ws.cell(row=total_row, column=1, value='Grand Total (Rs)')
        total_label.font = bold_font
        total_label.fill = total_fill
        total_label.alignment = right
        total_label.border = border

        total_value = ws.cell(row=total_row, column=7, value=values['grand_total'])
        total_value.font = bold_font
        total_value.fill = total_fill
        total_value.alignment = right
        total_value.number_format = '#,##0.00'
        total_value.border = border

        # Amount in Words
        words_row = total_row + 1
        ws.merge_cells(f'A{words_row}:G{words_row}')
        words_cell = ws.cell(
            row=words_row,
            column=1,
            value=f"Amount in Words: {values['amount_in_words']}"
        )
        words_cell.font = Font(name='Arial', bold=True, size=10, italic=True)
        words_cell.alignment = left

        # =========================
        # SAVE FILE
        # =========================
        output = io.BytesIO()
        wb.save(output)
        xlsx_data = output.getvalue()

        # Cleanup buffers
        for buf in image_buffers:
            buf.close()
        output.close()

        return xlsx_data